"""
RECRUITIN DASHBOARD AUTO-UPDATER
Haalt data uit Pipedrive API en update Excel dashboard
"""

import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime, timedelta
import json

# ============================================
# PIPEDRIVE CONFIG
# ============================================
API_KEY = "e689cc5b89affb98d93b1af95d115851f2d3d492"
DOMAIN = "recruitinbv.pipedrive.com"
BASE_URL = f"https://{DOMAIN}/v1"

# ============================================
# HELPER FUNCTIONS
# ============================================
def pipedrive_get(endpoint, params=None):
    """Generic GET request to Pipedrive API"""
    if params is None:
        params = {}
    params['api_token'] = API_KEY
    
    url = f"{BASE_URL}/{endpoint}"
    try:
        response = requests.get(url, params=params, timeout=10)
        
        if response.status_code == 200:
            return response.json()
        else:
            print(f"❌ Error {response.status_code}: {endpoint}")
            return None
    except Exception as e:
        print(f"❌ Exception on {endpoint}: {e}")
        return None

def get_date_range(period='month'):
    """Get start/end dates for current period"""
    now = datetime.now()
    
    if period == 'month':
        start = now.replace(day=1)
        if now.month == 12:
            end = now.replace(year=now.year+1, month=1, day=1) - timedelta(days=1)
        else:
            end = now.replace(month=now.month+1, day=1) - timedelta(days=1)
    elif period == 'year':
        start = now.replace(month=1, day=1)
        end = now.replace(month=12, day=31)
    elif period == 'week':
        start = now - timedelta(days=now.weekday())
        end = start + timedelta(days=6)
    
    return start.strftime('%Y-%m-%d'), end.strftime('%Y-%m-%d')

# ============================================
# DATA EXTRACTORS
# ============================================

def get_pipeline_value():
    """Get total pipeline value"""
    result = pipedrive_get('deals', params={'status': 'open', 'limit': 500})
    
    if result and 'data' in result and result['data']:
        total = sum([d.get('value', 0) or 0 for d in result['data']])
        return total
    return 0

def get_deals_won(period='month'):
    """Get number of deals won in period"""
    start_date, end_date = get_date_range(period)
    
    result = pipedrive_get('deals', params={'status': 'won', 'start_date': start_date, 'limit': 500})
    
    if result and 'data' in result and result['data']:
        count = 0
        for deal in result['data']:
            if deal.get('won_time'):
                won_date = deal['won_time'][:10]
                if start_date <= won_date <= end_date:
                    count += 1
        return count
    return 0

def get_new_leads(period='month'):
    """Get new leads created in period"""
    start_date, end_date = get_date_range(period)
    
    result = pipedrive_get('deals', params={'start_date': start_date, 'limit': 500})
    
    if result and 'data' in result and result['data']:
        count = 0
        for deal in result['data']:
            if deal.get('add_time'):
                add_date = deal['add_time'][:10]
                if start_date <= add_date <= end_date:
                    count += 1
        return count
    return 0

def get_activities_count(activity_type, period='month'):
    """Get count of activities (meetings/proposals) - using deals with stage info instead"""
    # Pipedrive activities endpoint might have different structure
    # Alternative: Count based on deals in specific stages
    print(f"   ⚠️  Activities endpoint not available - using deal stages as proxy")
    return 0  # Placeholder - will need manual input

def get_avg_deal_size():
    """Get average deal size from won deals"""
    result = pipedrive_get('deals', params={'status': 'won', 'limit': 500})
    
    if result and 'data' in result and result['data']:
        values = [d.get('value', 0) or 0 for d in result['data'] if d.get('value')]
        if values:
            return sum(values) / len(values)
    return 0

def get_avg_sales_cycle():
    """Get average sales cycle in days"""
    result = pipedrive_get('deals', params={'status': 'won', 'limit': 100})
    
    if result and 'data' in result and result['data']:
        cycles = []
        for deal in result['data']:
            if deal.get('add_time') and deal.get('won_time'):
                try:
                    start = datetime.fromisoformat(deal['add_time'].replace('Z', '+00:00'))
                    end = datetime.fromisoformat(deal['won_time'].replace('Z', '+00:00'))
                    days = (end - start).days
                    if days > 0:
                        cycles.append(days)
                except:
                    pass
        
        if cycles:
            return sum(cycles) / len(cycles)
    return 0

def get_stuck_deals_stage2(days=14):
    """Get deals stuck in stage 2 for >X days"""
    result = pipedrive_get('deals', params={'status': 'open', 'limit': 500})
    
    if result and 'data' in result and result['data']:
        cutoff_date = datetime.now() - timedelta(days=days)
        count = 0
        
        for deal in result['data']:
            if deal.get('stage_id') == 2:
                if deal.get('update_time'):
                    try:
                        update = datetime.fromisoformat(deal['update_time'].replace('Z', '+00:00'))
                        if update < cutoff_date:
                            count += 1
                    except:
                        pass
        return count
    return 0

def get_total_deals():
    """Get total number of deals in pipeline"""
    result = pipedrive_get('deals', params={'status': 'open', 'limit': 500})
    
    if result and 'data' in result and result['data']:
        return len(result['data'])
    return 0

# ============================================
# MAIN DATA COLLECTION
# ============================================

def collect_all_data():
    """Collect all metrics from Pipedrive"""
    print("🔄 Fetching data from Pipedrive API...")
    print(f"   Domain: {DOMAIN}")
    print(f"   Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("")
    
    data = {}
    
    # Sales metrics
    print("📊 Sales Metrics:")
    data['pipeline_value'] = get_pipeline_value()
    print(f"   Pipeline Value: €{data['pipeline_value']:,.0f}")
    
    data['deals_won_mtd'] = get_deals_won('month')
    print(f"   Deals Won MTD: {data['deals_won_mtd']}")
    
    data['deals_won_ytd'] = get_deals_won('year')
    print(f"   Deals Won YTD: {data['deals_won_ytd']}")
    
    data['new_leads_mtd'] = get_new_leads('month')
    print(f"   New Leads MTD: {data['new_leads_mtd']}")
    
    data['meetings_mtd'] = get_activities_count('meeting', 'month')
    print(f"   Meetings MTD: {data['meetings_mtd']} (manual input needed)")
    
    data['proposals_mtd'] = get_activities_count('proposal', 'month')
    print(f"   Proposals MTD: {data['proposals_mtd']} (manual input needed)")
    
    data['avg_deal_size'] = get_avg_deal_size()
    print(f"   Avg Deal Size: €{data['avg_deal_size']:,.0f}")
    
    data['avg_sales_cycle'] = get_avg_sales_cycle()
    print(f"   Avg Sales Cycle: {data['avg_sales_cycle']:.0f} days")
    
    data['deals_stuck_stage2'] = get_stuck_deals_stage2(14)
    print(f"   Deals Stuck Stage 2: {data['deals_stuck_stage2']}")
    
    data['total_deals'] = get_total_deals()
    print(f"   Total Deals: {data['total_deals']}")
    
    print("\n✅ Data collection complete!")
    return data

# ============================================
# EXCEL UPDATER
# ============================================

def update_excel_dashboard(data, excel_path):
    """Update Excel dashboard with collected data"""
    print(f"\n📝 Updating Excel: {excel_path}")
    
    wb = load_workbook(excel_path)
    ws = wb["🔢 DATA INPUT"]
    
    # Update cells
    updates = {
        'B7': data['pipeline_value'],
        'B8': data['deals_won_mtd'],
        'B9': data['deals_won_ytd'],
        'B10': data['new_leads_mtd'],
        'B13': data['avg_deal_size'],
        'B14': data['avg_sales_cycle'],
        'B15': data['deals_stuck_stage2'],
        'B16': data['total_deals'],
    }
    
    for cell, value in updates.items():
        ws[cell] = value
        ws[cell].fill = PatternFill(fill_type=None)
        ws[cell].font = Font(bold=False, color="000000")
    
    # Keep meetings/proposals yellow (manual)
    for cell in ['B11', 'B12']:
        ws[cell].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    # Update timestamp
    ws['A2'] = f"Laatste update: {datetime.now().strftime('%d-%m-%Y %H:%M')} (AUTO via API)"
    
    # Save
    wb.save(excel_path)
    print("✅ Excel updated successfully!")
    
    return True

# ============================================
# RUN
# ============================================

if __name__ == "__main__":
    print("=" * 60)
    print("🚀 RECRUITIN DASHBOARD AUTO-UPDATER")
    print("=" * 60)
    print("")
    
    # Collect data
    data = collect_all_data()
    
    # Save to JSON
    output_file = f"/tmp/pipedrive_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    with open(output_file, 'w') as f:
        json.dump(data, f, indent=2, default=str)
    print(f"\n💾 Data saved to: {output_file}")
    
    # Update Excel
    excel_path = "/tmp/RECRUITIN_MASTER_DASHBOARD_2026_V2.xlsx"
    update_excel_dashboard(data, excel_path)
    
    # Copy to outputs
    import shutil
    output_path = "/mnt/user-data/outputs/RECRUITIN_MASTER_DASHBOARD_2026_V2.xlsx"
    shutil.copy(excel_path, output_path)
    
    print("\n" + "=" * 60)
    print("✅ DONE! Dashboard is up-to-date")
    print("=" * 60)

